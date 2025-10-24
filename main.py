import streamlit as st
import pandas as pd
import plotly.express as px
import json
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

st.set_page_config(page_title="Simple Finance App", page_icon="💸", layout="wide")

category_file = "categories.json"
transactions_file = "transactions_data.json"
master_excel_file = "master_finance_tracker.xlsx"

# Load categories
if "categories" not in st.session_state:
    st.session_state.categories = {
        "Uncategorized": [],
    }

if os.path.exists(category_file):
    with open(category_file, "r") as f:
        st.session_state.categories = json.load(f)

# Load transactions data from previous sessions
if "transactions_df" not in st.session_state:
    if os.path.exists(transactions_file):
        try:
            transactions_data = pd.read_json(transactions_file)
            st.session_state.transactions_df = transactions_data
            st.session_state.data_loaded = True
        except:
            st.session_state.transactions_df = pd.DataFrame()
            st.session_state.data_loaded = False
    else:
        st.session_state.transactions_df = pd.DataFrame()
        st.session_state.data_loaded = False

def save_categories():
    with open(category_file, "w") as f:
        json.dump(st.session_state.categories, f)

def save_transactions():
    """Save current transactions to file for persistence"""
    if hasattr(st.session_state, 'transactions_df') and not st.session_state.transactions_df.empty:
        st.session_state.transactions_df.to_json(transactions_file, orient='records', date_format='iso')
        return True
    return False

def append_to_persistent_data():
    """Append current session data to persistent storage"""
    if not hasattr(st.session_state, 'current_session_df') or st.session_state.current_session_df.empty:
        st.warning("No current session data to append.")
        return False
    
    if st.session_state.transactions_df.empty:
        # No existing data, just save current session
        st.session_state.transactions_df = st.session_state.current_session_df.copy()
    else:
        # Merge with existing data and remove duplicates
        merged_df = pd.concat([st.session_state.transactions_df, st.session_state.current_session_df], ignore_index=True)
        merged_df = merged_df.drop_duplicates(subset=['Date', 'Merchant', 'Amount'], keep='last')
        st.session_state.transactions_df = merged_df
    
    save_transactions()
    return True

def create_master_excel():
    """Create or update the master Excel file with monthly summary"""
    if not hasattr(st.session_state, 'transactions_df') or st.session_state.transactions_df.empty:
        st.warning("No transaction data available to export.")
        return
    
    # Get current month/year
    current_date = datetime.now()
    current_month = current_date.strftime("%Y-%m")
    
    # Calculate monthly totals
    df = st.session_state.transactions_df.copy()
    
    # Ensure Date column is datetime
    if "Date" in df.columns:
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
        df["Month"] = df["Date"].dt.to_period('M').astype(str)
    else:
        df["Month"] = current_month
    
    # Handle separate Inflow/Outflow columns or legacy Amount column
    if "Inflow" in df.columns and "Outflow" in df.columns:
        # New format with separate columns
        outflow_df = df[["Month", "Outflow"]].copy()
        inflow_df = df[["Month", "Inflow"]].copy()
        
        # Group by month
        monthly_outflow = outflow_df.groupby("Month")["Outflow"].sum().reset_index()
        monthly_inflow = inflow_df.groupby("Month")["Inflow"].sum().reset_index()
        
        # Rename columns for consistency
        monthly_outflow = monthly_outflow.rename(columns={"Outflow": "Amount"})
        monthly_inflow = monthly_inflow.rename(columns={"Inflow": "Amount"})
        
    else:
        # Legacy format - treat all amounts as outflows
        if "Amount" not in df.columns:
            st.error("❌ No Amount column found in the data. Cannot create master Excel file.")
            return
            
        outflow_df = df.copy()
        outflow_df["Amount"] = outflow_df["Amount"].abs()  # Make positive for display
        
        # No inflow for now
        inflow_df = pd.DataFrame(columns=["Month", "Amount"])  # Empty dataframe
        
        # Group by month
        monthly_outflow = outflow_df.groupby("Month")["Amount"].sum().reset_index()
        monthly_inflow = inflow_df.groupby("Month")["Amount"].sum().reset_index()
    
    # Create master summary
    master_data = []
    
    # Get all unique months
    all_months = sorted(set(monthly_outflow["Month"].tolist() + monthly_inflow["Month"].tolist()))
    
    for month in all_months:
        outflow_amount = monthly_outflow[monthly_outflow["Month"] == month]["Amount"].sum()
        inflow_amount = monthly_inflow[monthly_inflow["Month"] == month]["Amount"].sum()
        net_amount = inflow_amount - outflow_amount
        
        master_data.append({
            "Month": month,
            "Outflow": outflow_amount,
            "Inflow": inflow_amount,
            "Net": net_amount
        })
    
    # Add current month if not exists
    if current_month not in [row["Month"] for row in master_data]:
        outflow_amount = monthly_outflow[monthly_outflow["Month"] == current_month]["Amount"].sum()
        inflow_amount = monthly_inflow[monthly_inflow["Month"] == current_month]["Amount"].sum()
        net_amount = inflow_amount - outflow_amount
        
        master_data.append({
            "Month": current_month,
            "Outflow": outflow_amount,
            "Inflow": inflow_amount,
            "Net": net_amount
        })
    
    # Create DataFrame
    master_df = pd.DataFrame(master_data)
    master_df = master_df.sort_values("Month")
    
    # Create Excel file with multiple sheets
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create Master Summary sheet
    ws_summary = wb.create_sheet("Master Summary")
    
    # Add headers
    headers = ["Month", "Outflow (CAD)", "Inflow (CAD)", "Net (CAD)"]
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    for row_idx, row in master_df.iterrows():
        ws_summary.cell(row=row_idx+2, column=1, value=row["Month"])
        ws_summary.cell(row=row_idx+2, column=2, value=row["Outflow"])
        ws_summary.cell(row=row_idx+2, column=3, value=row["Inflow"])
        ws_summary.cell(row=row_idx+2, column=4, value=row["Net"])
    
    # Format numbers
    for row in range(2, len(master_df) + 2):
        for col in range(2, 5):
            cell = ws_summary.cell(row=row, column=col)
            cell.number_format = '"$"#,##0.00'
    
    # Create Transaction Details sheet
    ws_transactions = wb.create_sheet("Transaction Details")
    
    # Add transaction data
    transaction_df = st.session_state.transactions_df.copy()
    transaction_df = transaction_df.sort_values("Date", ascending=False)
    
    for r in dataframe_to_rows(transaction_df, index=False, header=True):
        ws_transactions.append(r)
    
    # Format transaction sheet headers
    for cell in ws_transactions[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Auto-adjust column widths
    for ws in [ws_summary, ws_transactions]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the file
    wb.save(master_excel_file)
    return master_df

def categorize_transactions(df):
    df["Category"] = "Uncategorized"
    
    for category, keywords in st.session_state.categories.items():
        if category == "Uncategorized" or not keywords:
            continue
        
        lowered_keywords = [keyword.lower().strip() for keyword in keywords]
        
        for idx, row in df.iterrows():
            merchant = str(row["Merchant"]).lower().strip()
            if any(keyword in merchant for keyword in lowered_keywords):
                df.at[idx, "Category"] = category
                
    return df

def load_transactions(file):
    try:
        df = pd.read_csv(file)
        
        # Convert Date column to datetime if it exists
        if "Date" in df.columns:
            df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
        
        # Handle different CSV formats
        if "Inflow" in df.columns and "Outflow" in df.columns:
            # New format with separate Inflow/Outflow columns
            pass  # Keep as is
        elif "Amount" in df.columns:
            # Legacy format - convert to separate columns
            df["Inflow"] = df[df["Amount"] > 0]["Amount"].fillna(0)
            df["Outflow"] = df[df["Amount"] < 0]["Amount"].abs().fillna(0)
        else:
            st.error("Unsupported CSV format. Expected columns: Date, Description, Inflow, Outflow")
            return None
            
        return categorize_transactions(df)
    except Exception as e:
        st.error(f"Error processing file: {str(e)}")
        return None

def add_keyword_to_category(category, keyword):
    keyword = keyword.strip()
    if keyword and keyword not in st.session_state.categories[category]:
        st.session_state.categories[category].append(keyword)
        save_categories()
        return True
    
    return False        

def main():
    st.title("Simple Finance Dashboard")
    
    # Show data status
    if st.session_state.data_loaded:
        st.info(f"📁 Loaded {len(st.session_state.transactions_df)} transactions from previous sessions")
    
    uploaded_file = st.file_uploader("Upload your transaction CSV file", type=["csv"])

    df = pd.DataFrame()  # Default empty DataFrame

    if uploaded_file is not None:
        # create a simple token that changes when the uploaded file changes
        upload_token = f"{uploaded_file.name}:{getattr(uploaded_file, 'size', None)}"
        should_load = (
            "current_session_df" not in st.session_state
            or st.session_state.current_session_df.empty
            or st.session_state.get("upload_token") != upload_token
        )

        if should_load:
            df = load_transactions(uploaded_file)
            if df is not None:
                st.session_state.current_session_df = df.copy()
                st.session_state.upload_token = upload_token
                st.success(f"✅ Loaded {len(df)} transactions for this session")

    # Show tabs including new Master Tracker tab
    tab1, tab2, tab3 = st.tabs(["💸 Outflow", "💰 Inflow", "📊 Master Tracker"])

    with tab1:
        st.subheader("💸 Outflow Transactions (Expenses)")
        
        new_category = st.text_input("New Category Name")
        add_button = st.button("Add Category")
        
        if add_button and new_category:
            if new_category not in st.session_state.categories:
                st.session_state.categories[new_category] = []
                save_categories()
                st.success(f"✅ Added category: {new_category}")
                # Don't auto-rerun, let user click Apply Changes manually

        # Show current session data for editing
        # Use session state data if available, otherwise use current df
        display_df = st.session_state.current_session_df if hasattr(st.session_state, 'current_session_df') and not st.session_state.current_session_df.empty else df
        
        if not display_df.empty:
            # Filter for outflow transactions only
            if "Inflow" in display_df.columns and "Outflow" in display_df.columns:
                outflow_df = display_df[display_df["Outflow"] > 0].copy()
                
                if not outflow_df.empty:
                    st.write(f"📊 Found {len(outflow_df)} outflow transactions")
                    
                    # Show outflow summary
                    total_outflow = outflow_df["Outflow"].sum()
                    st.metric("Total Outflow", f"${total_outflow:,.2f}")
                    
                    # Add sorting controls
                    col_sort1, col_sort2 = st.columns([1, 3])
                    with col_sort1:
                        sort_by = st.selectbox("Sort by:", ["None", "Category", "Date", "Outflow", "Merchant"])
                    with col_sort2:
                        sort_order = st.selectbox("Order:", ["Ascending", "Descending"])
                    
                    # Apply sorting
                    outflow_df_sorted = outflow_df.copy()
                    if sort_by != "None":
                        ascending = sort_order == "Ascending"
                        if sort_by in outflow_df_sorted.columns:
                            outflow_df_sorted = outflow_df_sorted.sort_values(by=sort_by, ascending=ascending)

                    # Only use columns that exist
                    if "Date" in outflow_df_sorted.columns:
                        outflow_df_sorted["Date"] = pd.to_datetime(outflow_df_sorted["Date"], errors="coerce")
                    
                    display_cols = [col for col in ["Date", "Description", "Merchant", "Outflow", "Category"] if col in outflow_df_sorted.columns]
                else:
                    st.info("No outflow transactions found in current session data.")
                    display_cols = []
            else:
                # Legacy format - show all transactions as outflow
                st.write("💡 **Edit like Excel**: Click any cell to edit directly!")
                
                # Add sorting controls
                col_sort1, col_sort2 = st.columns([1, 3])
                with col_sort1:
                    sort_by = st.selectbox("Sort by:", ["None", "Category", "Date", "Amount", "Merchant"])
                with col_sort2:
                    sort_order = st.selectbox("Order:", ["Ascending", "Descending"])
                
                # Apply sorting
                df_sorted = display_df.copy()
                if sort_by != "None":
                    ascending = sort_order == "Ascending"
                    if sort_by in df_sorted.columns:
                        df_sorted = df_sorted.sort_values(by=sort_by, ascending=ascending)

                # Only use columns that exist
                if "Date" in df_sorted.columns:
                    df_sorted["Date"] = pd.to_datetime(df_sorted["Date"], errors="coerce")
                
                display_cols = [col for col in ["Date", "Description", "Merchant", "Amount", "Category"] if col in df_sorted.columns]
                outflow_df_sorted = df_sorted

            if display_cols:  # Only show editor if there are transactions
                # Add a "Delete" column
                outflow_df_sorted["Delete"] = False
                display_cols.append("Delete")

                # Ensure Date column is datetime for editing
                if "Date" in outflow_df_sorted.columns:
                    outflow_df_sorted["Date"] = pd.to_datetime(outflow_df_sorted["Date"], errors="coerce")

                # Excel-like editing with better column configuration
                column_config = {
                    "Description": st.column_config.TextColumn(
                        "Description",
                        help="Click to edit description"
                    ),
                    "Merchant": st.column_config.TextColumn(
                        "Merchant",
                        help="Click to edit merchant name"
                    ),
                    "Category": st.column_config.SelectboxColumn(
                        "Category",
                        options=list(st.session_state.categories.keys()),
                        help="Click to change category"
                    ),
                    "Delete": st.column_config.CheckboxColumn(
                        "Delete",
                        help="Check to delete this transaction"
                    )
                }
                
                # Add Date column config only if Date column exists and is datetime
                if "Date" in outflow_df_sorted.columns and pd.api.types.is_datetime64_any_dtype(outflow_df_sorted["Date"]):
                    column_config["Date"] = st.column_config.DateColumn(
                        "Date", 
                        format="DD/MM/YYYY",
                        help="Click to edit date"
                    )
                
                # Add amount column configurations based on format
                if "Outflow" in display_cols:
                    column_config["Outflow"] = st.column_config.NumberColumn(
                        "Outflow", 
                        format="%.2f CAD",
                        help="Click to edit outflow amount",
                        min_value=0.0,
                        step=0.01
                    )
                elif "Amount" in display_cols:
                    column_config["Amount"] = st.column_config.NumberColumn(
                        "Amount", 
                        format="%.2f CAD",
                        help="Click to edit amount. For splitting: divide by number of people",
                        min_value=0.0,
                        step=0.01
                    )
                
                edited_df = st.data_editor(
                    outflow_df_sorted[display_cols],
                    column_config=column_config,
                    hide_index=True,
                    use_container_width=True,
                    key="outflow_editor"
                )

                # Action buttons
                col1, col2, col3, col4 = st.columns(4)
                
                with col1:
                    save_button = st.button("💾 Apply Changes", type="primary", use_container_width=True)
                
                with col2:
                    if st.button("🔄 Reset to Original", use_container_width=True):
                        st.rerun()
                
                with col3:
                    if st.button("📊 Show Outflow Summary", use_container_width=True):
                        st.session_state.show_outflow_summary = True
                
                with col4:
                    append_button = st.button("📁 Append to Master Data", type="secondary", use_container_width=True)

                if save_button:
                    try:
                        # 1) Start from the edited grid as the source of truth
                        edited_df_clean = edited_df.copy()

                        # 2) Separate deletions and keepers using the checkbox column
                        to_delete_mask = edited_df_clean["Delete"] == True
                        keep_df = edited_df_clean.loc[~to_delete_mask].drop(columns=["Delete"]).copy()

                        # 3) If you have Outflow/Amount columns, ensure numeric types
                        for col in ["Outflow", "Amount"]:
                            if col in keep_df.columns:
                                keep_df[col] = pd.to_numeric(keep_df[col], errors="coerce").fillna(0.0)

                        # 4) (Optional) add changed merchants into category keyword lists
                        #    Here we simply go row-by-row and ensure any (Merchant, Category) pairs are learned.
                        if "Merchant" in keep_df.columns and "Category" in keep_df.columns:
                            for _, row in keep_df.iterrows():
                                merchant = str(row.get("Merchant", "")).strip()
                                cat = row.get("Category", "Uncategorized")
                                if merchant and cat in st.session_state.categories:
                                    add_keyword_to_category(cat, merchant)

                        # 5) Rebuild the full current_session_df:
                        source_df = st.session_state.current_session_df.copy()
                        if "Inflow" in source_df.columns and "Outflow" in source_df.columns:
                            df_without_outflow = source_df[source_df["Outflow"] <= 0].copy()
                            updated_df = pd.concat([df_without_outflow, keep_df], ignore_index=True)

                        else:
                            # Legacy format: keep_df already represents the edited set
                            updated_df = keep_df.copy()

                        # 6) Store back to session
                        st.session_state.current_session_df = updated_df

                        # 7) User feedback
                        deleted_count = int(to_delete_mask.sum())
                        if deleted_count > 0:
                            st.success(f"✅ Deleted {deleted_count} transaction(s)")
                        st.success("✅ Changes applied successfully!")

                        st.rerun()

                    except Exception as e:
                        st.error(f"❌ Error applying changes: {str(e)}")

                if append_button:
                    try:
                        if append_to_persistent_data():
                            st.success("✅ Current session data appended to master database!")
                            st.balloons()
                            # Clear current session data
                            st.session_state.current_session_df = pd.DataFrame()
                            st.rerun()
                        else:
                            st.error("❌ Failed to append data.")
                    except Exception as e:
                        st.error(f"❌ Error appending data: {str(e)}")

                # Show outflow summary if requested
                if st.session_state.get("show_outflow_summary", False) or save_button:
                    st.subheader('📊 Outflow Summary (Current Session)')
                    
                    # Use the current session data for summary
                    summary_df = st.session_state.current_session_df if hasattr(st.session_state, 'current_session_df') and not st.session_state.current_session_df.empty else df
                    
                    # Check if we have the required columns
                    if "Category" not in summary_df.columns:
                        st.error("❌ No Category column found. Please ensure your data has been categorized.")
                        return
                    
                    if "Outflow" in summary_df.columns:
                        # Filter for outflow transactions only
                        outflow_data = summary_df[summary_df["Outflow"] > 0].copy()
                        
                        if not outflow_data.empty:
                            category_totals = outflow_data.groupby("Category")["Outflow"].sum().reset_index()
                            category_totals = category_totals.sort_values("Outflow", ascending=False)
                            
                            grand_total = category_totals["Outflow"].sum()
                            
                            # Append grand total as a new row
                            grand_total_row = pd.DataFrame([{"Category": "Total", "Outflow": grand_total}])
                            category_totals_with_total = pd.concat([category_totals, grand_total_row], ignore_index=True)

                            # Display summary
                            st.dataframe(
                                category_totals_with_total, 
                                column_config={
                                 "Outflow": st.column_config.NumberColumn("Outflow", format="%.2f CAD")   
                                },
                                use_container_width=True,
                                hide_index=True
                            )
                            
                            # Pie chart
                            fig = px.pie(
                                category_totals,
                                values="Outflow",
                                names="Category",
                                title="Outflow by Category (Current Session)"
                            )
                            st.plotly_chart(fig, use_container_width=True)
                        else:
                            st.info("No outflow transactions found.")
                    elif "Amount" in summary_df.columns:
                        # Legacy format
                        category_totals = summary_df.groupby("Category")["Amount"].sum().reset_index()
                        category_totals = category_totals.sort_values("Amount", ascending=False)
                        
                        grand_total = category_totals["Amount"].sum()

                        # Append grand total as a new row
                        grand_total_row = pd.DataFrame([{"Category": "Total", "Amount": grand_total}])
                        category_totals_with_total = pd.concat([category_totals, grand_total_row], ignore_index=True)

                        # Display summary
                        st.dataframe(
                            category_totals_with_total, 
                            column_config={
                             "Amount": st.column_config.NumberColumn("Amount", format="%.2f CAD")   
                            },
                            use_container_width=True,
                            hide_index=True
                        )
                        
                        # Pie chart
                        fig = px.pie(
                            category_totals,
                            values="Amount",
                            names="Category",
                            title="Expenses by Category (Current Session)"
                        )
                        st.plotly_chart(fig, use_container_width=True)
                    else:
                        st.error("❌ No Outflow or Amount column found in the data.")
        else:
            st.info("Upload a transaction file to see outflow data.")

    with tab2:
        st.subheader("💰 Inflow Transactions (Income)")
        
        # Use session state data if available, otherwise use current df
        display_df_inflow = st.session_state.current_session_df if hasattr(st.session_state, 'current_session_df') and not st.session_state.current_session_df.empty else df
        
        if not display_df_inflow.empty:
            # Filter for inflow transactions only
            if "Inflow" in display_df_inflow.columns and "Outflow" in display_df_inflow.columns:
                inflow_df = display_df_inflow[display_df_inflow["Inflow"] > 0].copy()
                
                if not inflow_df.empty:
                    st.write(f"📊 Found {len(inflow_df)} inflow transactions")
                    
                    # Show inflow summary
                    total_inflow = inflow_df["Inflow"].sum()
                    st.metric("Total Inflow", f"${total_inflow:,.2f}")
                    
                    # Display inflow transactions
                    display_cols = [col for col in ["Date", "Description", "Merchant", "Inflow", "Category"] if col in inflow_df.columns]
                    
                    # Add delete column
                    inflow_df["Delete"] = False
                    display_cols.append("Delete")
                    
                    # Ensure Date column is datetime for editing
                    if "Date" in inflow_df.columns:
                        inflow_df["Date"] = pd.to_datetime(inflow_df["Date"], errors="coerce")

                    # Excel-like editing for inflow
                    column_config = {
                        "Description": st.column_config.TextColumn("Description"),
                        "Merchant": st.column_config.TextColumn("Merchant"),
                        "Inflow": st.column_config.NumberColumn("Inflow", format="%.2f CAD", min_value=0.0, step=0.01),
                        "Category": st.column_config.SelectboxColumn("Category", options=list(st.session_state.categories.keys())),
                        "Delete": st.column_config.CheckboxColumn("Delete")
                    }
                    
                    # Add Date column config only if Date column exists and is datetime
                    if "Date" in inflow_df.columns and pd.api.types.is_datetime64_any_dtype(inflow_df["Date"]):
                        column_config["Date"] = st.column_config.DateColumn("Date", format="DD/MM/YYYY")
                    
                    edited_inflow_df = st.data_editor(
                        inflow_df[display_cols],
                        column_config=column_config,
                        hide_index=True,
                        use_container_width=True,
                        key="inflow_editor"
                    )
                    
                    # Action buttons for inflow
                    col1, col2 = st.columns(2)
                    with col1:
                        if st.button("💾 Save Inflow Changes", type="primary"):
                            try:
                                # Simple approach: Create new dataframe with only non-deleted rows
                                rows_to_keep = []
                                
                                for i, (edited_idx, row) in enumerate(edited_inflow_df.iterrows()):
                                    if row["Delete"] != True:  # Keep non-deleted rows
                                        # Get the corresponding row from the inflow dataframe
                                        if i < len(inflow_df):
                                            original_row = inflow_df.iloc[i].copy()
                                            
                                            # Update the original row with any changes from the edited row
                                            # Update inflow amount
                                            original_row["Inflow"] = row.get("Inflow", 0)
                                            
                                            # Update category
                                            original_row["Category"] = row.get("Category", original_row.get("Category", "Uncategorized"))
                                            
                                            # Add merchant to category keywords if category changed
                                            merchant = str(row.get("Merchant", "")).strip()
                                            old_category = inflow_df.iloc[i].get("Category", "Uncategorized")
                                            new_category = row.get("Category", old_category)
                                            if merchant and new_category != old_category:
                                                add_keyword_to_category(new_category, merchant)
                                            
                                            rows_to_keep.append(original_row)
                                
                                # Create new dataframe with only the rows we want to keep
                                if rows_to_keep:
                                    new_inflow_df = pd.DataFrame(rows_to_keep).reset_index(drop=True)
                                    
                                    # Count how many were deleted
                                    deleted_count = len(inflow_df) - len(rows_to_keep)
                                    
                                    # Update the main dataframe - remove old inflow transactions and add updated ones
                                    # Remove all inflow transactions first
                                    df_without_inflow = display_df_inflow[display_df_inflow["Inflow"] <= 0].copy()
                                    # Add back the updated inflow transactions
                                    updated_df_inflow = pd.concat([df_without_inflow, new_inflow_df], ignore_index=True)
                                    
                                    # Update session state
                                    st.session_state.current_session_df = updated_df_inflow.copy()
                                    
                                    if deleted_count > 0:
                                        st.success(f"✅ Deleted {deleted_count} inflow transaction(s)")
                                    st.success("✅ Inflow changes applied successfully!")
                                else:
                                    st.warning("⚠️ All inflow transactions were deleted!")
                                
                                st.rerun()
                                
                            except Exception as e:
                                st.error(f"❌ Error applying inflow changes: {str(e)}")
                                st.write("Please try again or refresh the page.")
                    
                    with col2:
                        if st.button("📊 Inflow Summary"):
                            # Show inflow by category
                            inflow_by_category = inflow_df.groupby("Category")["Inflow"].sum().reset_index()
                            inflow_by_category = inflow_by_category.sort_values("Inflow", ascending=False)
                            
                            st.subheader("💰 Inflow by Category")
                            st.dataframe(
                                inflow_by_category,
                                column_config={"Inflow": st.column_config.NumberColumn("Inflow", format="%.2f CAD")},
                                use_container_width=True,
                                hide_index=True
                            )
                else:
                    st.info("No inflow transactions found in current session data.")
            else:
                st.info("Inflow/Outflow columns not found. Upload a file with separate Inflow/Outflow columns.")
        else:
            st.info("Upload a transaction file to see inflow data.")

    with tab3:
        st.subheader("📊 Master Finance Tracker")
        st.write("Track your monthly inflow, outflow, and net amounts over time.")
        
        if hasattr(st.session_state, 'transactions_df') and not st.session_state.transactions_df.empty:
            # Show current monthly summary
            current_date = datetime.now()
            current_month = current_date.strftime("%Y-%m")
            
            df_master = st.session_state.transactions_df.copy()
            if "Date" in df_master.columns:
                df_master["Date"] = pd.to_datetime(df_master["Date"], errors="coerce")
                df_master["Month"] = df_master["Date"].dt.to_period('M').astype(str)
            else:
                df_master["Month"] = current_month
            
            # Calculate totals for all months using separate Inflow/Outflow columns
            if "Inflow" in df_master.columns and "Outflow" in df_master.columns:
                all_outflow = df_master["Outflow"].sum()
                all_inflow = df_master["Inflow"].sum()
                all_net = all_inflow - all_outflow
                
                # Calculate current month totals
                current_month_data = df_master[df_master["Month"] == current_month]
                current_outflow = current_month_data["Outflow"].sum()
                current_inflow = current_month_data["Inflow"].sum()
                current_net = current_inflow - current_outflow
            else:
                # Fallback for legacy Amount column
                all_outflow = df_master["Amount"].abs().sum()
                all_inflow = 0
                all_net = all_inflow - all_outflow
                
                current_month_data = df_master[df_master["Month"] == current_month]
                current_outflow = current_month_data["Amount"].abs().sum()
                current_inflow = 0
                current_net = current_inflow - current_outflow
            
            # Display summary
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Current Month", current_month)
                st.caption(f"All-time: {len(df_master)} transactions")
            with col2:
                st.metric("Outflow (Current)", f"${current_outflow:,.2f}")
                st.caption(f"All-time: ${all_outflow:,.2f}")
            with col3:
                st.metric("Inflow (Current)", f"${current_inflow:,.2f}")
                st.caption(f"All-time: ${all_inflow:,.2f}")
            with col4:
                st.metric("Net (Current)", f"${current_net:,.2f}", delta=f"{current_net:+,.2f}")
                st.caption(f"All-time: ${all_net:,.2f}")
            
            # Create and download Excel file
            if st.button("📥 Export Master Excel File", type="primary"):
                with st.spinner("Creating master Excel file..."):
                    master_df = create_master_excel()
                    
                    # Display the master summary
                    st.subheader("Master Monthly Summary")
                    st.dataframe(
                        master_df,
                        column_config={
                            "Month": "Month",
                            "Outflow": st.column_config.NumberColumn("Outflow (CAD)", format="$%.2f"),
                            "Inflow": st.column_config.NumberColumn("Inflow (CAD)", format="$%.2f"),
                            "Net": st.column_config.NumberColumn("Net (CAD)", format="$%.2f")
                        },
                        use_container_width=True,
                        hide_index=True
                    )
                    
                    # Provide download link
                    if os.path.exists(master_excel_file):
                        with open(master_excel_file, "rb") as file:
                            st.download_button(
                                label="📥 Download Master Excel File",
                                data=file.read(),
                                file_name=master_excel_file,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        st.success("✅ Master Excel file created successfully!")
        else:
            st.info("Upload transaction data and append to master database to see your finance tracker.")

main()   